import openpyxl
import datetime
import csv

myData = 'MyDatabase.xlsx'
wb = openpyxl.load_workbook(myData)
ws1 = wb['Courses']
ws2 = wb['Sections']

data = 'data.xlsx'
wb2 = openpyxl.load_workbook(data)
dataSheet = wb2['Courses']

class Course:

    def __init__(self, courseCode):
        self.__code = courseCode
        # self.__name = name
        # self.__sectionList = section      #This whould be a list of section object (aggregation)
        # self.__examDates = examDate
        # self.__examTime = examTime
        self.__sectionList = []

        for code,name,examDate,examTime in ws1.rows:
            if code.value == courseCode:
                self.__name = name.value
                self.__examDates = examDate.value
                self.__examTime = examTime.value
                self.__sectionList = []

        for i in range(2,ws2.max_row+1):
            l = []
            if ws2.cell(row=i,column=1).value == courseCode:
                for j in range(1,3):
                    d = ws2.cell(row=i,column=j).value
                    l.append(d)
                if(len(l) != 0):
                    section = Section(l[0],l[1])
                    self.__sectionList.append(section)

    def get_code(self):
        return self.__code

    def get_sections(self):
        return self.__sectionList

    def print_all_sections(self):
        for i in self.__sectionList:
            print(i)

    def get_info(self):
        l = {
            "code":self.__code,
            "name":self.__name,
            "section":self.__sectionList,
            "examDate":self.__examDates,
            "examTime":self.__examTime}
        return l

    def populate_section(self, type, number, password):
        tryPass = input("Enter the admin password: ")
        if tryPass == password:
            for i in range (number):
                counter = 1
                for i in self.__sectionList:
                    if (i.get_info()["sectionNo"][0] == type):
                        counter+=1
                sectionNumber = type[0] + str(counter)
                nHour = int(input("Enter the number of hours: "))
                hourList = []
                hourStr = 0
                hour = int(input("Enter the starting Hour: "))
                for i in range(nHour):
                    hourList.append(hour)
                    hourStr = hourStr*10 + hour
                    hour+=1
                nDays = int(input("Enter the number of days in the section: "))
                daysList = []
                daysStr = ''
                for j in range(nDays):
                    day = input("Enter the day " + str(j+1) + " for " + type + " " + str(i+1) +": ")
                    daysList.append(day)
                    daysStr += day + ','
                venue = int(input("Enter the venue for " + type + " " + str(i+1) +": "))
                nProff = int(input("Enter the number of professors in the section: "))
                proffList = []
                proffStr = ''
                for j in range(nProff):
                    proff = input("Enter the professor " + str(j+1) + "for" + type + " " + str(i+1) +": ")
                    proffList.append(proff)
                    proffStr += proff + ','
                daysStr = daysStr[0:-1]
                proffStr = proffStr[0:-1]
                row = [self.__code, sectionNumber, hourStr, daysStr, venue, proffStr]
                ws2.append(row)
                wb.save(myData)
                section = Section(self.__code, sectionNumber)
                self.__sectionList.append(section)
                print("Section has been added successfully.")
        else:
            print("\nWrong Pin.\n")

class Section():

    def __init__(self, code, sectionNo):
        self.__code = code
        self.__sectionNo = sectionNo
        for code1, sectionNumber, hour, day, venue, proff in ws2.rows:
            if code1.value == code and sectionNumber.value == sectionNo :
                self.__timing = hour.value
                self.__day = day.value.split(',')
                self.__venue = venue.value
                self.__proff = proff.value.split(',')

    def get_info(self):
        data = {"code":self.__code,
                "sectionNo":self.__sectionNo,
                "timing":self.__timing,
                "day":self.__day,
                "venue":self.__venue,
                "proff":self.__proff}
        return data

    def __str__(self):
        return '[' + str(self.__code) + ',' + str(self.__sectionNo) + ',' + str(self.__timing) + ',' + str(self.__day) + ',' + str(self.__venue) + ',' + str(self.__proff) + ']'

class TimeTable:

    def __init__(self):
        self.table = [[0]*9]*6
        self.__listOfCourses = []

    def enroll_subject(self,course,section):
        courseDetails = course.get_info()
        sectionDetails = section.get_info()
        code = courseDetails["code"]
        days = sectionDetails["day"]
        week = {'M':0 , 'T':1 , 'W':2 , 'Th':3 , 'F':4 , 'S':5}
        for i in self.__listOfCourses:
            if i.get_info()["examDate"] == courseDetails["examDate"] and i.get_info()["examTime"] == courseDetails["examTime"]:
                print("\nClash of Examination. " + i.get_info()["name"] + " has examination at the same time.\n")
        for i in range(0,len(days)):
            days[i] = week[days[i]]
        hours = sectionDetails["timing"]
        indices = []
        a = int(hours)
        while(a>0):
            hour = a%10 - 1
            a = a//10
            for day in days:
                if self.table[day][hour]==0:
                    print(hour,end=',')
                    print(day)
                    indices.append([day,hour])
                else:
                    print("\nClash Between Courses. " + str(self.table[i][j]) + " already at this slot.\n")
                    return
        # self.table[3][0] = courseDetails["code"]
        for day,hour in indices:
            print('inside')
            print(day,end=',')
            print(hour)
            self.table[day][hour] = courseDetails["code"]
            print(self.table)
        print("\nSuccessfully Enrolled.\n")
        self.__listOfCourses.append(course)
        self.export_to_csv('timetable.csv')

    def withdraw_course(self,course):
        for i in range(0,6):
            for j in range(0,9):
                # print('i = ' + str(i))
                if self.table[i][j] == course.get_code():
                    self.table[i][j] = 0
                    print("You have sucessfully withdrawn from the course")
                    return
        print("You were not enrolled into this course.")

    def show_courses(self):
        for i in self.__listOfCourses:
            print(i.get_info())

    def get_courses(self):
        return self.__listOfCourses

    def print_timetable(self):
        rows = []
        with open('timetable.csv', 'r') as csvfile:
            csvReader = csv.reader(csvfile)
            for row in csvReader:
                rows.append(row)

            for row in rows:
                for col in row:
                    print("%10s"%col,end=" ")
                print('\n')

    def export_to_csv(self,file):
        with open(file, 'w') as csvfile:
            csvWriter = csv.writer(csvfile)
            fields = [1,2,3,4,5,6,7,8,9]
            csvWriter.writerow(fields)
            # week = ['M','T','W','Th','F','S']
            # j=0
            for i in self.table:
                # i.append(week[j])
                csvWriter.writerow(i)
                # j+=1

def populate_course(password):
    tryPass = input("Enter the admin password: ")
    if tryPass == password:
        counter=1
        availableCourses = []    # List of the courses that are not added to personal database
        print("The available courses are: ")
        # iterating throgh all the rows in the database
        for i in range (2,dataSheet.max_row+1):
            flag=1
            code = dataSheet.cell(row=i,column=2).value
            #checking if the course already exists in personal database
            for j in range(2,ws1.max_row+1):
                if(ws1.cell(row=j,column=1).value == code):
                    flag=0
            #displaying all the available courses
            if (flag == 1):
                availableCourses.append(code)
                print(str(counter) + ' ' + code)
                counter+=1
        choice = int(input("Enter Your Choice: "))
        # selecting the required row from the database and inserting it to the personal database
        for i in range (2,dataSheet.max_row+1):
            code = dataSheet.cell(row=i,column=2).value
            if(code == availableCourses[choice-1]):
                # making the row to be appended
                row = []
                for j in range(2,dataSheet.max_column+1):
                    row.append(dataSheet.cell(row=i,column=j).value)
                ws1.append(row)
                wb.save(myData)
    else:
        print("Wrong Pin.")

def get_all_courses():
    l=[]
    for i in range(2,ws1.max_row+1):
        a = ws1.cell(row=i,column=1)
        l.append(a.value)
    return l

def menu(tt):
    print('Welcome to the Time Table Manager.')
    password = input("Please Set your admin Password: ")
    print("\nYour admin password has been set.\n")
    choice = 0
    while(choice != 11):
        nav = ('''
    1.  Add Courses to the database.
    2.  Add Section(s) to a Course.
    3.  Enroll to a Course.
    4.  Withdraw Form a Course.
    5.  See list of enrolled Courses.
    6.  See List of all the sections of a given Course.
    7.  See Details of a Course.
    8.  See Details of a Section of a Course.
    9.  See Current Time Table.
    10. Get Time Table as CSV File.
    11. Exit
    ''')
        print(nav)
        choice = int(input("Enter your choice: "))
        
        if(choice == 1):
            populate_course(password)

        elif(choice == 2):
            print("These are the Courses: ")
            l = get_all_courses()
            print(l)
            code = input("Which Course do you want to add a section in: ")
            course1 = Course(code)
            sectionType = input("What type of section do you want to add?\nL for Lecture\nT for Tutorial\nP for Laboratory. ")
            nSection = int(input("How many sections do you want to add? "))
            course1.populate_section(sectionType,nSection,password)

        elif(choice == 3):
            lCourse = get_all_courses()
            print(lCourse)
            courseCode = input("Enter the Code of the course you want to enroll into: ")
            course = Course(courseCode)
            course.print_all_sections()
            sectionCode = input("Enter the Section Code of the section you want to Enroll into: ")
            section = Section(courseCode,sectionCode)
            tt.enroll_subject(course,section)

        elif(choice == 4):
            lCourse = tt.get_courses()
            if(len(lCourse) == 0):
                print("You are not enrolled in any course currently.")
            else:
                for i in lCourse:
                    print(i.get_code())
                courseCode = input("Enter the Code of the course you want to enroll into: ")
                course = Course(courseCode)
                tt.withdraw_course(course)
        
        elif(choice == 5):
            print("These are the list of all the courses you have enrolled in: ")
            lCourse = tt.get_courses()
            if(len(lCourse) == 0):
                print("You are not enrolled in any course currently.")
            else:
                print(lCourse)

        elif(choice == 6):
            lCourse = get_all_courses()
            print(lCourse)
            courseCode = input("Enter the code of the course you want to see the sections of: ")
            course1 = Course(courseCode)
            course1.print_all_sections()

        elif(choice == 7):
            lCourse = get_all_courses()
            print(lCourse)
            courseCode = input("Enter the code of the course you want to see the sections of: ")
            course1 = Course(courseCode)
            details = course1.get_info()
            print(details)

        elif(choice == 8):
            lCourse = get_all_courses()
            print(lCourse)
            courseCode = input("Enter the code of the course you want to see the sections of: ")
            course1 = Course(courseCode)
            course1.print_all_sections()
            sectionCode = input("Enter the Section Code of the section you want to Enroll into: ")
            section = Section(courseCode,sectionCode)
            details = section.get_info()
            print(details)

        elif(choice == 9):
            tt.print_timetable()

        elif(choice == 10):
            fileName = input("Enter the file name in which you want to save the timetable: ")
            fileName = fileName + '.csv'
            file = open(fileName,'a')
            tt.export_to_csv(file)

tt = TimeTable()
menu(tt)
fileName = 'timetable.csv'
f = open(fileName, "w+")
f.close()